import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from itertools import combinations, permutations
import matplotlib.pyplot as plt
import random
def get_average(list):
    sum = 0
    for item in list:
        sum += item
    return sum/len(list)

def get_std_variance(list):
    sum = 0
    average = get_average(list)
    for item in list:
        sum += (item - average)**2
    std_dev = (sum/(len(list)-1))**0.5
    return std_dev
workbook = load_workbook('C:\\Users\\zengyangzhou\\Desktop\\工作\\AI4\\SharePrices.xlsx')
sheets = workbook.get_sheet_names()  # 从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[0])
rows = booksheet.rows
columns = booksheet.columns
result = open('C:\\Users\\zengyangzhou\\Desktop\\工作\\AI4\\ass_4_v2.txt', 'w', encoding='utf-8')
for row in rows:
    line = [col.value for col in row]
#通过坐标读取值
data1 = []
data2 = []
data3 = []
data4 = []
data5 = []
data6 = []
data7 = []
data8 = []
data9 = []
data10 = []
for i in range(2, 501):
    #cell_11 = booksheet.cell(row=i, column=1).value
    data1.append(booksheet.cell(row=i, column=1).value)
    data2.append(booksheet.cell(row=i, column=2).value)
    data3.append(booksheet.cell(row=i, column=3).value)
    data4.append(booksheet.cell(row=i, column=4).value)
    data5.append(booksheet.cell(row=i, column=5).value)
    data6.append(booksheet.cell(row=i, column=6).value)
    data7.append(booksheet.cell(row=i, column=7).value)
    data8.append(booksheet.cell(row=i, column=8).value)
    data9.append(booksheet.cell(row=i, column=9).value)
    data10.append(booksheet.cell(row=i, column=10).value)
    #print(cell_11)
#print(data)
data = [[], data1, data2, data3, data4, data5, data6, data7, data8, data9, data10]
#总的收益率
R1 = (data1[-1] - data1[0]) / data1[0]
R2 = (data2[-1] - data2[0]) / data2[0]
R3 = (data3[-1] - data3[0]) / data3[0]
R4 = (data4[-1] - data4[0]) / data4[0]
R5 = (data5[-1] - data5[0]) / data5[0]
R6 = (data6[-1] - data6[0]) / data6[0]
R7 = (data7[-1] - data7[0]) / data7[0]
R8 = (data8[-1] - data8[0]) / data8[0]
R9 = (data9[-1] - data9[0]) / data9[0]
R10 = (data10[-1] - data10[0]) / data10[0]
Return = [0, R1, R2, R3, R4, R5, R6, R7, R8, R9, R10]
#计算每只股票的标准差
r1 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data1):
        break
    r = (data1[i+1] - data1[i]) / data1[i]
    r1.append(r)
r1_stdev = get_std_variance(r1)

r2 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data2):
        break
    r = (data2[i+1] - data2[i]) / data2[i]
    r2.append(r)
r2_stdev = get_std_variance(r1)

r3 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data3):
        break
    r = (data3[i+1] - data3[i]) / data3[i]
    r3.append(r)
r3_stdev = get_std_variance(r1)

r4 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data4):
        break
    r = (data4[i+1] - data4[i]) / data4[i]
    r4.append(r)
r4_stdev = get_std_variance(r1)

r5 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data5):
        break
    r = (data5[i+1] - data5[i]) / data5[i]
    r5.append(r)
r5_stdev = get_std_variance(r1)

r6 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data6):
        break
    r = (data6[i+1] - data6[i]) / data6[i]
    r6.append(r)
r6_stdev = get_std_variance(r1)

r7 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data7):
        break
    r = (data7[i+1] - data7[i]) / data7[i]
    r7.append(r)
r7_stdev = get_std_variance(r1)

r8 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data8):
        break
    r = (data8[i+1] - data8[i]) / data8[i]
    r8.append(r)
r8_stdev = get_std_variance(r1)

r9 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data9):
        break
    r = (data9[i+1] - data9[i]) / data9[i]
    r9.append(r)
r9_stdev = get_std_variance(r1)

r10 = []
for i in range(2, 501):
    i = i - 2
    if i+1 == len(data10):
        break
    r = (data10[i+1] - data10[i]) / data10[i]
    r10.append(r)
r10_stdev = get_std_variance(r1)

stock_order = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
r_stdev = [0, r1_stdev, r2_stdev, r3_stdev, r4_stdev, r5_stdev, r6_stdev, r7_stdev, r8_stdev, r9_stdev, r10_stdev]
stock_name = {1: 'HSBA', 2: 'BARC', 3: 'BA', 4: 'VOD', 5: 'BATS', 6: 'TSCO', 7: 'PRU', 8: 'AVIVA', 9: 'GSK', 10: 'BP'}
risk_profit = {}
f1 = plt.figure(1)

def gen_ran(C):
    ran_list = []
    for i in range(C):
        ran_list.append(random.randint(0, 10))
    ran_list = ran_list/(np.sum(ran_list))
    #print(ran_list)
    return ran_list
max_ratio = 0
for mont_ca in range(20):
    for C in range(1, 11):    #选多少支股票
        #W = [0] * C    #初始化权重
        W = gen_ran(C)
        for comb in list(combinations(stock_order, C)):
            Profit = 0
            Risk = 0
            ind1 = 0
            for stock in comb:  #C = 3, 【1，2, 3】
                Wei1 = W[ind1]
                ind1 += 1
                Profit = Profit + Wei1 * Return[stock]
                ind2 = 0
                for stock2 in comb:
                    Wei2 = W[ind2]
                    ind2 += 1
                    Risk = Risk + Wei1 * Wei2 * np.cov(data[stock], data[stock2])[0][1]

            Risk = np.sqrt(Risk)
            plt.scatter(Risk, Profit, alpha=0.6)
            sharpe_ratio = Profit / Risk
            if sharpe_ratio > max_ratio:
                max_ratio = sharpe_ratio
                result.write('\n')
                result.write('when we chose stock(s) of \t', )
                for x, y in enumerate(comb):
                    result.write(str(stock_name[y]))
                    result.write('\t')
                    result.write('W=')
                    result.write(str(W[x]))
                    result.write('\t')
                result.write('The Return is \t')
                result.write(str(Profit))
                result.write('\t')
                result.write(' and Risk is \t')
                result.write(str(Risk))
                result.write('\t')
                result.write(' and Sharpe Ratio \t')
                result.write(str(max_ratio))
                result.write('\n')
plt.show()
